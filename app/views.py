from flask import request, redirect, render_template, url_for
from app import app
import locale
locale.setlocale(locale.LC_TIME, "sp") # swedish
import openpyxl
from datetime import datetime, timedelta
import calendar
#import numpy

def cerradas():
    #File Log
    FILEPATH_LOG = open(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Log_Cerradas.txt','a')

    #Read Excel
    NAME_FILE=openpyxl.load_workbook(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Prueba.xlsx')
    sheet =  NAME_FILE['Cerradas']
    DIA_ONE = 1
    DIA_TWO = 2

    filepath_cerrradas = r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Cerradas.xlsx'
    wb = openpyxl.Workbook()
    wb.save(filepath_cerrradas)

    FILEPATH_Cerradas = openpyxl.load_workbook(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Cerradas.xlsx')

    sheet_Cerradas = FILEPATH_Cerradas.active
    FINAL_COUNT_NUM_TOTAL_ROW = 2

    for r in range(1, 1 + 1):
        for c in range(1, 24):
            d = sheet.cell(row=r, column=c)
            #print('%-8s' % d.value, end='')
            #print('', end=""),
            row_final = sheet_Cerradas.cell(row=r, column=c)
            row_final.value = d.value
        #print('')


    count_num_total_rows = 1

    next = False

    #my_date_actual_compare_with_excel = (datetime.now()- timedelta(2)).strftime('%Y-%m-%d')
    my_date_actual_compare_with_excel = datetime.now()
    FILEPATH_LOG.write('-----------------------------------\n')
    FILEPATH_LOG.write(str(datetime.now())+'\n')
    my_date_yesterday_compare_with_excel = (datetime.now()- timedelta(0)).strftime('%Y-%m-%d')
    my_year_actual =  datetime.now().strftime('%Y-%m-%d')

    #print ("my_date_actual_compare_with_excel : ", my_date_actual_compare_with_excel)
    #print ("my_date_yesterday_compare_with_excel: ", my_date_yesterday_compare_with_excel)

    #We have the total rows
    while(next == False):
        column_name_f = str("f"+str(count_num_total_rows))
        if (sheet[column_name_f].value == None):
            next = True
        else:
            count_num_total_rows = count_num_total_rows + 1


    day_studying_number = int(my_date_actual_compare_with_excel.strftime('%d'))

    #print (day_studying_number)

    day_studying = my_date_actual_compare_with_excel.weekday()
    day_studying_number_change_month = int(my_date_actual_compare_with_excel.strftime('%d'))

    month_actual_compare_change = 0
    month_actual_compare_change_less = 0

    month_actual_compare_change = my_date_actual_compare_with_excel.strftime('%m')
    month_actual_compare_change_less = my_date_yesterday_compare_with_excel[5:7]

    #print ("MONTH_actual_compare_change_LESS: ", month_actual_compare_change_less)

    my_year_actual = my_year_actual[0:4]

    if(day_studying_number_change_month == 1 and day_studying == 1):
        # SI HAY CAMBIO DE MES Y ES MARTES 1; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        #print ("****************SI HAY CAMBIO DE MES Y ES MARTES 1***********************")
        FILEPATH_LOG.write("SI HAY CAMBIO DE MES Y ES MARTES 1\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before or day_actual_excel == last_day_month_before-1 or day_actual_excel == last_day_month_before-2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before or day_actual_excel == last_day_month_before - 1 or day_actual_excel == last_day_month_before - 2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before or day_actual_excel == last_day_month_before - 1 or day_actual_excel == last_day_month_before - 2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIDO DE MES Y ES MARTES 1

    elif(day_studying_number_change_month == 2 and day_studying == 1):
        # SI HAY CAMBIO DE MES Y ES MARTES 2; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO
        #print("**************SI HAY CAMBIO DE MES Y ES MARTES 2******************")
        FILEPATH_LOG.write("SI HAY CAMBIO DE MES Y ES MARTES 2\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less) - 1))
        last_day_month_before = int(last_day_month_before[1])
        #print(last_day_month_before)

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((compare_month == 0 and day_actual_excel == DIA_ONE) or (compare_month == 1 and (day_actual_excel == last_day_month_before  or day_actual_excel == last_day_month_before - 1))):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((compare_month == 0 and day_actual_excel == DIA_ONE) or (compare_month == 1 and (day_actual_excel == last_day_month_before  or day_actual_excel == last_day_month_before - 1))):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((compare_month == 0 and day_actual_excel == DIA_ONE) or (compare_month == 1 and (day_actual_excel == last_day_month_before or day_actual_excel == last_day_month_before - 1))):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIO DE MES Y ES MARTES 2

    elif(day_studying_number_change_month == 3 and day_studying == 1):
        # SI HAY CAMBIO DE MES Y ES MARTES 3; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO
        FILEPATH_LOG.write("SI HAY CAMBIO DE MES Y ES MARTES 3\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less) - 1))
        last_day_month_before = int(last_day_month_before[1])
        #print(last_day_month_before)

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 0 and (day_actual_excel == DIA_ONE or day_actual_excel == DIA_TWO)) or (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 0 and (day_actual_excel == DIA_ONE or day_actual_excel == DIA_TWO)) or (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 0 and (day_actual_excel == DIA_ONE or day_actual_excel == DIA_TWO)) or (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIO DE MES Y ES MARTES 3

    elif(day_studying_number_change_month == 1 and not day_studying == 0 and not day_studying == 1):
        # SI HAY CAMBIDO DE MES 1 Y NO ES LUNES NI MARTES;

        #print ("************SI HAY CAMBIDO DE MES 1 Y NO ES LUNES NI MARTES********************")
        FILEPATH_LOG.write("SI HAY CAMBIDO DE MES 1 Y NO ES LUNES NI MARTES\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIDO DE MES 1 Y NO ES LUNES NI MARTES;

    elif(day_studying_number_change_month == 1 and day_studying == 0):
        # SI HAY CAMBIDO DE MES Y ES LUNES 1; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        #print ("************SI HAY CAMBIDO DE MES Y ES LUNES 1***********************")
        FILEPATH_LOG.write("SI HAY CAMBIDO DE MES Y ES LUNES 1")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)

                    if (compare_month == 1 and (day_actual_excel == last_day_month_before-2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                print('%-8s' % d.value, end='')
                                print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before - 2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before - 2)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIDO DE MES Y ES LUNES 1;

    elif(day_studying_number_change_month == 2 and day_studying == 0):
        # SI HAY CAMBIDO DE MES Y ES LUNES 2; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        #print ("*******************SI HAY CAMBIDO DE MES Y ES LUNES 2********************")
        FILEPATH_LOG.write("SI HAY CAMBIDO DE MES Y ES LUNES 2")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before-1)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before - 1)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before - 1)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIDO DE MES Y ES LUNES 2;

    elif(day_studying_number_change_month == 3 and day_studying == 0):
        # SI HAY CAMBIDO DE MES Y ES LUNES 3; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        #print ("***************SI HAY CAMBIDO DE MES Y ES LUNES 3*******************")
        FILEPATH_LOG.write("SI HAY CAMBIDO DE MES Y ES LUNES 3\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if (compare_month == 1 and (day_actual_excel == last_day_month_before )):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    # FIN SI HAY CAMBIDO DE MES Y ES LUNES 3;

    elif(not day_studying_number_change_month == 3 and  not day_studying_number_change_month == 2 and not day_studying_number_change_month == 1 and day_studying == 0):
        # SI ES UN LUNES CUALQUIERA; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        FILEPATH_LOG.write("SI ES UN LUNES CUALQUIERA\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less)-1))
        last_day_month_before = int(last_day_month_before[1])
        #print (last_day_month_before )

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 3) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 3) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        #print("SELECCIONAMOS TIWS")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 3) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    #  SI ES UN MARTES CUALQUIERA;

    elif (not day_studying_number_change_month == 3 and not day_studying_number_change_month == 2 and not day_studying_number_change_month == 1 and day_studying == 1):
        # SI ES UN MARTES CUALQUIERA; LA COMPROBACION DE CAMBIO DE MES LA HACEMOS EN UN IF DE ABAJO

        FILEPATH_LOG.write("SI ES UN MARTES CUALQUIERA\n")
        last_day_month_before = calendar.monthrange(int(my_year_actual), int(int(month_actual_compare_change_less) - 1))
        last_day_month_before = int(last_day_month_before[1])
        #print(last_day_month_before)

        # We have the files that we are interested
        for final_count_num_total_rows in range(1, count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 1
                         or day_actual_excel_compare - day_actual_excel == 2
                         or day_actual_excel_compare - day_actual_excel == 3)
                            and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 1
                         or day_actual_excel_compare - day_actual_excel == 2
                         or day_actual_excel_compare - day_actual_excel == 3)
                            and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print(day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    compare_month = int(month_actual_excel_compare) - int(month_actual_excel)
                    if ((day_actual_excel_compare - day_actual_excel == 1
                         or day_actual_excel_compare - day_actual_excel == 2
                         or day_actual_excel_compare - day_actual_excel == 3)
                            and (int(month_actual_excel_compare) - int(month_actual_excel) == 0)):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    #  SI ES UN LUNES CUALQUIERA;

    elif(day_studying == 2 or day_studying == 3 or day_studying == 4):
        #DE MIERCOLES A VIERNES SIN CAMBIO DE MES; CASO MAS FACIL
        FILEPATH_LOG.write("DE MIERCOLES A VIERNES SIN CAMBIO DE MES\n")
        #We have the files that we are interested
        for final_count_num_total_rows in range(1,count_num_total_rows):
            column_name_f = str("f" + str(final_count_num_total_rows))
            column_name_k = str("k" + str(final_count_num_total_rows))

            if (sheet[column_name_f].value == 'TIWS' or sheet[column_name_f].value == 'TIWS '):
                #print (column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print (day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    #print("Month Actual Compare", int(month_actual_excel_compare) - int(month_actual_excel) )
                    if((day_actual_excel_compare - day_actual_excel == 1) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0) ):
                        FILEPATH_LOG.write("SELECCIONAMOS TIWS\n")
                        for r in range(final_count_num_total_rows , final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW , column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TISA ' or sheet[column_name_f].value == 'TISA'):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print (day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    #print("Month Actual Compare", int(month_actual_excel_compare) - int(month_actual_excel) )
                    if((day_actual_excel_compare - day_actual_excel == 1) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0) ):
                        FILEPATH_LOG.write("SELECCIONAMOS TISA\n")
                        for r in range(final_count_num_total_rows , final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW , column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

            elif (sheet[column_name_f].value == 'TEDIG' or sheet[column_name_f].value == 'TEDIG '):
                #print(column_name_f)
                if (sheet[column_name_k].value) != 'OPEN':
                    cadena = str(sheet[column_name_k].value)
                    day_actual_excel = cadena[8:10]
                    month_actual_excel = cadena[5:7]
                    day_actual_excel = int(day_actual_excel)
                    #print (day_actual_excel)
                    #print("Month Actual", month_actual_excel)
                    day_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%d')
                    month_actual_excel_compare = my_date_actual_compare_with_excel.strftime('%m')
                    day_actual_excel_compare = int(day_actual_excel_compare)
                    month_actual_excel_compare = int(month_actual_excel_compare)
                    #print(day_actual_excel_compare)
                    #print("Month Actual Compare", int(month_actual_excel_compare) - int(month_actual_excel) )
                    if((day_actual_excel_compare - day_actual_excel == 1) and (int(month_actual_excel_compare) - int(month_actual_excel) == 0) ):
                        FILEPATH_LOG.write("SELECCIONAMOS TEDIG\n")
                        for r in range(final_count_num_total_rows , final_count_num_total_rows + 1):
                            for c in range(1, 24):
                                d = sheet.cell(row=r, column=c)
                                #print('%-8s' % d.value, end='')
                                #print('', end=""),
                                row_final = sheet_Cerradas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW , column=c)
                                row_final.value = d.value
                            #print('')
                        FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Cerradas.save(filepath_cerrradas)
    #FIN DE MARTES A VIERNES SIN CAMBIO DE MES

    else:
        FILEPATH_LOG.write("NO ENTRO EN NIGUNO DE LAS OPCIONES\n")

    FILEPATH_LOG.write('-----------------------------------\n')
    FILEPATH_LOG.write("\n")
    FILEPATH_LOG.close()
    print("Terminado Cerradas\n")





def abiertas():
    # File Log
    FILEPATH_LOG = open(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Log_Abiertas.txt', 'a')

    # Read Excel
    NAME_FILE = openpyxl.load_workbook(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Prueba.xlsx')
    sheet = NAME_FILE['Abiertas']

    filepath_abiertas = r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Abiertas.xlsx'
    wb = openpyxl.Workbook()
    wb.save(filepath_abiertas)

    FILEPATH_Abiertas = openpyxl.load_workbook(r'C:\Users\usr1CR\.PyCharmCE2018.2\proyects\probando_jinja2\probando_jinja2\excel\Abiertas.xlsx')

    sheet_Abiertas = FILEPATH_Abiertas.active
    FINAL_COUNT_NUM_TOTAL_ROW = 2

    for r in range(1, 1 + 1):
        for c in range(1, 24):
            d = sheet.cell(row=r, column=c)
            row_final = sheet_Abiertas.cell(row=r, column=c)
            row_final.value = d.value

    count_num_total_rows = 1

    next = False

    FILEPATH_LOG.write('-----------------------------------\n')
    FILEPATH_LOG.write(str(datetime.now()) + '\n')

    # We have the total rows
    while (next == False):
        column_name_h = str("h" + str(count_num_total_rows))
        if (sheet[column_name_h].value == None):
            next = True
        else:
            count_num_total_rows = count_num_total_rows + 1

    # We have the files that we are interested
    for final_count_num_total_rows in range(1, count_num_total_rows):
        column_name_h = str("h" + str(final_count_num_total_rows))
        column_name_m = str("m" + str(final_count_num_total_rows))

        if (sheet[column_name_h].value == 'TIWS' or sheet[column_name_h].value == 'TIWS '
                or sheet[column_name_h].value == 'TEDIG' or sheet[column_name_h].value == 'TEDIG '
                or sheet[column_name_h].value == 'TISA' or sheet[column_name_h].value == 'TISA '):

            if ((sheet[column_name_m].value) == 'OPEN' or (sheet[column_name_m].value == 'OPEN ')):
                FILEPATH_LOG.write("SELECCIONAMOS TIWS o TEDIG o TISA\n")
                for r in range(final_count_num_total_rows, final_count_num_total_rows + 1):
                    for c in range(1, 24):
                        d = sheet.cell(row=r, column=c)
                        row_final = sheet_Abiertas.cell(row=FINAL_COUNT_NUM_TOTAL_ROW, column=c)
                        row_final.value = d.value
                FINAL_COUNT_NUM_TOTAL_ROW = FINAL_COUNT_NUM_TOTAL_ROW + 1

        FILEPATH_Abiertas.save(filepath_abiertas)

    FILEPATH_LOG.write('-----------------------------------\n')
    FILEPATH_LOG.write("\n")
    FILEPATH_LOG.close()
    print("Terminado Abiertas\n")


my_date=datetime.now()

month=""
if   (my_date.strftime('%m') == '01'):
    month = "Enero"
elif (my_date.strftime('%m') == '02'):
    month = "Febrero"
elif (my_date.strftime('%m') == '03'):
    month = "Marzo"
elif (my_date.strftime('%m') == '04'):
    month = "Arbil"
elif (my_date.strftime('%m') == '05'):
    month = "Mayo"
elif (my_date.strftime('%m') == '06'):
    month = "Junio"
elif (my_date.strftime('%m') == '07'):
    month = "Julio"
elif (my_date.strftime('%m') == '08'):
    month = "Agosto"
elif (my_date.strftime('%m') == '09'):
    month = "Septiembre"
elif (my_date.strftime('%m') == '10'):
    month = "Octubre"
elif (my_date.strftime('%m') == '11'):
    month = "Noviembre"
elif (my_date.strftime('%m') == '12'):
    month = "Diciembre"


day=""
if   (my_date.weekday()== 0):
    day = "Lunes"
elif (my_date.weekday() == 1):
    day = "Martes"
elif (my_date.weekday() == 2):
    day = "Miércoles"
elif (my_date.weekday() == 3):
    day = "Jueves"
elif (my_date.weekday() == 4):
    day= "Viernes"
elif (my_date.weekday() == 5):
    day = "Sábado"
elif (my_date.weekday() == 6):
    day = "Domingo"


cerradas()
abiertas()



'''
@app.route('/')

def index():
    return render_template('index.html', name_columns=['Infinity', 'Cisco SR', 'Cisco RMA', 'Ticket SMC', 'Cliente',
                                                       'Sala de apertura', 'Adm. de circuito', 'Salas afectadas',
                                                       'País','Fecha de cierre','Escalado','Proactiva','Responsable',
                                                       'Motivo de apertura','Resolución','Tiempo abierta',
                                                       'Fecha de apertura'],
                           month_actual=month,my_date=datetime.now(),day_actual=day)

'''